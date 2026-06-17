from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.utils.text import slugify
from datetime import date
from calendar import monthrange
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from ..models import Billing, BillingDetail, Company, AuditLog
from ..forms import BillingFilterForm, BillingGenerateForm
from django.template.loader import render_to_string
from ..utils.billing import generate_billing_for_company, calculate_estimated_cost, simulate_billing_for_company
from ..utils.system_settings import get_system_settings
from ..utils.decorators import super_admin_required


@login_required
def billing_list(request):
    """Lista de Cobranças"""
    user = request.user
    
    # Filtrar por permissão
    if user.is_super_admin():
        billings = Billing.objects.select_related('company', 'generated_by').all()
    elif user.is_admin_company() and user.company:
        billings = Billing.objects.filter(company=user.company)
    else:
        messages.error(request, _('Acesso negado.'))
        return redirect('dashboard')
    
    # Filtros
    form = BillingFilterForm(request.GET or None)
    
    if form.is_valid():
        company = form.cleaned_data.get('company')
        period_start = form.cleaned_data.get('period_start')
        period_end = form.cleaned_data.get('period_end')
        
        if company and user.is_super_admin():
            billings = billings.filter(company=company)
        
        if period_start:
            billings = billings.filter(period_start__gte=period_start)
        
        if period_end:
            billings = billings.filter(period_end__lte=period_end)
    
    # Ordenar por data decrescente
    billings = billings.order_by('-period_start', '-created_at')
    
    # Paginação
    paginator = Paginator(billings, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': _('Cobranças'),
        'billings': page_obj,
        'form': form,
    }
    
    return render(request, 'billing/list.html', context)


@login_required
@super_admin_required
def billing_generate(request):
    """Gerar Cobrança"""
    if request.method == 'POST':
        form = BillingGenerateForm(request.POST)
        if form.is_valid():
            period_start = form.cleaned_data['period_start']
            period_end = form.cleaned_data['period_end']
            companies = form.cleaned_data.get('companies')
            
            if not companies:
                companies = Company.objects.filter(status='active')
            
            success_count = 0
            error_count = 0
            error_details = []
            
            for company in companies:
                try:
                    existing = Billing.objects.filter(
                        company=company,
                        period_start=period_start,
                        period_end=period_end
                    ).exists()
                    
                    if existing:
                        error_count += 1
                        error_details.append(f'{company.name}: cobrança já existe para este período')
                        continue
                    
                    billing = generate_billing_for_company(
                        company=company,
                        period_start=period_start,
                        period_end=period_end,
                        generated_by=request.user
                    )
                    
                    if billing:
                        success_count += 1
                        
                        AuditLog.objects.create(
                            user=request.user,
                            action='billing_generated',
                            description=f'Cobrança gerada para {company.name} - Período: {period_start} a {period_end}',
                            ip_address=request.META.get('REMOTE_ADDR')
                        )
                    else:
                        error_count += 1
                        error_details.append(f'{company.name}: sem valores a cobrar neste período')
                
                except Exception as e:
                    error_count += 1
                    error_details.append(f'{company.name}: {str(e)}')
                    continue
            
            msg = _('Geração concluída: {} sucesso, {} erros').format(success_count, error_count)
            if error_details:
                msg += '\n\n' + '\n'.join(f'• {d}' for d in error_details)
            messages.success(request, msg)
            return redirect('billing_list')
    else:
        form = BillingGenerateForm()
    
    context = {
        'page_title': _('Gerar Cobrança'),
        'form': form,
    }
    
    return render(request, 'billing/generate.html', context)


@login_required
@super_admin_required
def billing_preview(request):
    """Preview de Cobrança (simula sem persistir)"""
    today = date.today()
    initial = {
        'period_start': date(today.year, today.month, 1),
        'period_end': date(today.year, today.month, monthrange(today.year, today.month)[1]),
    }
    form = BillingGenerateForm(request.GET or None, initial=initial)
    preview_results = None
    totals = None
    
    if form.is_valid():
        period_start = form.cleaned_data['period_start']
        period_end = form.cleaned_data['period_end']
        companies = form.cleaned_data.get('companies')
        
        if not companies:
            companies = Company.objects.filter(status='active').order_by('name')
        
        preview_results = []
        for company in companies:
            sim = simulate_billing_for_company(company, period_start, period_end)
            preview_results.append({
                'company': company,
                'estimated_total': sim['total_value'],
                'active_members': sim['item_count'],
                'full_count': sim['full_count'],
                'proportional_count': sim['proportional_count'],
                'details': sim['details'],
            })
        
        from decimal import Decimal
        totals = {
            'active_members': sum(r['active_members'] for r in preview_results),
            'full_count': sum(r['full_count'] for r in preview_results),
            'proportional_count': sum(r['proportional_count'] for r in preview_results),
            'estimated_total': sum(r['estimated_total'] for r in preview_results),
        }
    
    context = {
        'page_title': _('Preview de Cobrança'),
        'form': form,
        'preview_results': preview_results,
        'totals': totals,
    }
    return render(request, 'billing/preview.html', context)


@login_required
def billing_detail(request, pk):
    """Detalhes da Cobrança"""
    from decimal import Decimal

    user = request.user
    billing = get_object_or_404(Billing, pk=pk)
    
    if user.is_admin_company() and user.company != billing.company:
        messages.error(request, _('Acesso negado.'))
        return redirect('billing_list')
    
    details = billing.details.select_related(
        'user', 'member', 'chatbot'
    ).order_by('member__name')
    
    full_details = [d for d in details if d.billing_type == 'full']
    prop_details = [d for d in details if d.billing_type == 'proportional']
    full_count = len(full_details)
    prop_count = len(prop_details)
    full_total = sum((d.value for d in full_details), Decimal('0.00'))
    prop_total = sum((d.value for d in prop_details), Decimal('0.00'))
    avg_days = round(sum(d.days_active for d in details) / max(len(details), 1), 1)
    
    if billing.company.billing_mode == 'per_user':
        admin_d = [d for d in details if d.user_id is not None]
        member_d = [d for d in details if d.user_id is None]
        by_type = {
            'admin': {'count': len(admin_d), 'total': sum((d.value for d in admin_d), Decimal('0.00'))},
            'member': {'count': len(member_d), 'total': sum((d.value for d in member_d), Decimal('0.00'))},
        }
        by_chatbot = []
        template = 'billing/detail_per_user.html'
    else:
        chatbot_groups = {}
        for d in details:
            name = d.chatbot.name if d.chatbot else 'Sem chatbot'
            if name not in chatbot_groups:
                chatbot_groups[name] = {'count': 0, 'total': Decimal('0.00')}
            chatbot_groups[name]['count'] += 1
            chatbot_groups[name]['total'] += d.value
        by_chatbot = [{'name': k, 'count': v['count'], 'total': v['total']} for k, v in chatbot_groups.items()]
        by_type = {}
        template = 'billing/detail_per_chatbot.html'
    
    summary_stats = {
        'total_items': len(details),
        'full_count': full_count,
        'proportional_count': prop_count,
        'full_total': full_total,
        'proportional_total': prop_total,
        'by_type': by_type,
        'by_chatbot': by_chatbot,
        'avg_days': avg_days,
    }
    
    type_chart_data = None
    chatbot_chart_data = None
    if by_type:
        type_chart_data = {
            'labels': [str(_('Admin')) if k == 'admin' else str(_('Membro')) for k in by_type.keys()],
            'values': [float(v['total']) for v in by_type.values()],
        }
    if by_chatbot:
        chatbot_chart_data = {
            'labels': [g['name'] for g in by_chatbot],
            'values': [float(g['total']) for g in by_chatbot],
        }
    
    context = {
        'page_title': _('Detalhes da Cobrança'),
        'billing': billing,
        'details': details,
        'summary_stats': summary_stats,
        'type_chart_data': type_chart_data,
        'chatbot_chart_data': chatbot_chart_data,
    }
    
    return render(request, template, context)


@login_required
def billing_export_csv(request, pk):
    """Exportar Cobrança em CSV"""
    user = request.user
    billing = get_object_or_404(Billing, pk=pk)
    
    # Verificar permissão
    if user.is_admin_company() and user.company != billing.company:
        messages.error(request, _('Acesso negado.'))
        return redirect('billing_list')
    
    # Gerar CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="cobranca_{billing.company.name}_{billing.period_start}_{billing.period_end}.csv"'
    
    # BOM para UTF-8
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Cabeçalho
    if billing.company.billing_mode == 'per_user':
        writer.writerow([
            'Usuário/Membro',
            'Tipo',
            'Preço Base',
            'Data de Ativação',
            'Tipo de Cobrança',
            'Valor'
        ])
        
        # Dados
        for detail in billing.details.select_related('user', 'member', 'chatbot').all():
            target = detail.user or detail.member
            target_type = 'Admin Empresa' if detail.user else 'Funcionário'
            
            writer.writerow([
                str(target),
                target_type,
                f"{billing.company.currency_symbol} {detail.unit_price:.2f}",
                detail.activation_date.strftime('%d/%m/%Y'),
                detail.get_billing_type_display(),
                f"{billing.company.currency_symbol} {detail.value:.2f}"
            ])
    else:
        writer.writerow([
            'Usuário/Membro',
            'Chatbot',
            'Preço Base',
            'Data de Ativação',
            'Tipo de Cobrança',
            'Valor'
        ])
        
        # Dados
        for detail in billing.details.select_related('user', 'member', 'chatbot').all():
            target = detail.user or detail.member
            
            writer.writerow([
                str(target),
                detail.chatbot.name if detail.chatbot else '-',
                f"{billing.company.currency_symbol} {detail.unit_price:.2f}",
                detail.activation_date.strftime('%d/%m/%Y'),
                detail.get_billing_type_display(),
                f"{billing.company.currency_symbol} {detail.value:.2f}"
            ])
    
    # Total
    writer.writerow([])
    writer.writerow(['Total', '', '', '', '', f"{billing.company.currency_symbol} {billing.total_value:.2f}"])
    
    return response


@login_required
def billing_export_excel(request, pk):
    """Exportar Cobrança em Excel"""
    user = request.user
    billing = get_object_or_404(Billing, pk=pk)

    if user.is_admin_company() and user.company != billing.company:
        messages.error(request, _('Acesso negado.'))
        return redirect('billing_list')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = str(_('Cobrança'))

    if billing.company.billing_mode == 'per_user':
        header = ['Usuário/Membro', 'Tipo', 'Preço Base', 'Data de Ativação', 'Tipo de Cobrança', 'Valor']
    else:
        header = ['Usuário/Membro', 'Chatbot', 'Preço Base', 'Data de Ativação', 'Tipo de Cobrança', 'Valor']
    sheet.append(header)

    details = billing.details.select_related('user', 'member', 'chatbot').order_by('member__name')
    for detail in details:
        target = detail.user or detail.member
        activation_str = detail.activation_date.strftime('%d/%m/%Y')
        row = [str(target)]

        if billing.company.billing_mode == 'per_user':
            row.append('Admin Empresa' if detail.user else 'Funcionário')
        else:
            row.append(detail.chatbot.name if detail.chatbot else '-')

        row.append(float(detail.unit_price))
        row.extend([activation_str, detail.get_billing_type_display(), float(detail.value)])
        sheet.append(row)

    sheet.append([])
    sheet.append(['Total', '', '', '', '', float(billing.total_value)])

    for idx in range(1, len(header) + 1):
        sheet.column_dimensions[get_column_letter(idx)].width = 25

    filename = f"cobranca_{slugify(billing.company.name)}_{billing.period_start}_{billing.period_end}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response

@login_required
def billing_export_pdf(request, pk):
    """Exportar Cobrança em PDF"""
    user = request.user
    billing = get_object_or_404(Billing, pk=pk)

    if user.is_admin_company() and user.company != billing.company:
        messages.error(request, _('Acesso negado.'))
        return redirect('billing_list')

    details = billing.details.select_related('user', 'member', 'chatbot').order_by('member__name')

    html = render_to_string('billing/invoice.html', {
        'billing': billing,
        'details': details,
        'generated_at': timezone.now(),
        'system_settings': get_system_settings(),
    })

    from weasyprint import HTML
    pdf = HTML(string=html).write_pdf()

    filename = f"cobranca_{slugify(billing.company.name)}_{billing.period_start}_{billing.period_end}.pdf"
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@super_admin_required
def billing_delete(request, pk):
    """Excluir cobrança"""
    billing = get_object_or_404(Billing, pk=pk)

    if request.method == 'POST':
        company_name = billing.company.name
        period_start = billing.period_start.strftime('%d/%m/%Y')
        period_end = billing.period_end.strftime('%d/%m/%Y')

        billing.delete()

        messages.success(
            request,
            _('Cobrança de {company} referente ao período {start} a {end} excluída com sucesso.').format(
                company=company_name,
                start=period_start,
                end=period_end
            )
        )
        return redirect('billing_list')

    context = {
        'page_title': _('Excluir Cobrança'),
        'billing': billing,
    }
    return render(request, 'billing/delete_confirm.html', context)
